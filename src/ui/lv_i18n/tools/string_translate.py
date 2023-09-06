#! python3
from openpyxl import load_workbook

wb = load_workbook('language_03.21.xlsx')           #多语言表格
ws = wb.active

wb1 = load_workbook('FD152_多国语言_211229.xlsx')   #翻译稿
ws1 = wb1.active

for row in range(5, ws.max_row + 1):                #按行遍历多语言表格
    color = ws.cell(row, 2).fill.fgColor.rgb    
    if color == "00000000":                         #如果此行的字符串ID单元格背景色是白色则不复制翻译内容
        continue
    str_id = ws.cell(row, 2).value
    if  str_id:
        for col in range(4, ws.max_column + 1):
            lang = ws.cell(3, col).value
            for col1 in range(6, ws1.max_column + 1):
                lang1 = ws1.cell(1, col1).value
                if lang == lang1:                               #判断语言类型一致后进行后续翻译copy
                    for row1 in range(3, ws1.max_row + 1):
                        str_id1 = ws1.cell(row1, 3).value
                        if str_id == str_id1:
                            text = ws1.cell(row1, col1).value
                            if text:
                                print(text)
                                ws.cell(row, col).value =text

wb.save('language_03.21.xlsx')