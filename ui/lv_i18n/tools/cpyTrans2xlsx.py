#! python3
import sys
import re
import os
import shutil
from openpyxl import load_workbook

args = sys.argv[:]
path = os.getcwd()
dic_file_dir = os.path.dirname(path)

wb = load_workbook('language_03.21.xlsx')
ws = wb.active

for folderName, subfolders, filenames in os.walk(dic_file_dir):             #遍历文件
    for filename in filenames:
        isyml = re.search(".yml", filename)                                 #筛选yml文件
        if isyml:
            f=open(dic_file_dir + "\\" + filename,'r+', encoding='UTF-8')
            flist=f.readlines()
            lang = flist[2]
            lang = lang.split("lang: ")
            lang = lang[1].split('\n')
            lang = lang[0].strip()                                          #截取语言类型
            print(lang)
            for col in range(3, ws.max_column + 1):                         #遍历excel表中的语言类型
                lang1 = ws.cell(2, col).value
                if lang == lang1:
                    print(lang)
                    total_lines = len(flist)
                    for line in range(3, total_lines):                      #逐行遍历字典文件
                        id = flist[line]
                        id = id.split(":")
                        id = id[0].strip()                                  #截取字典文件中的id
                        text = flist[line]
                        pos = 0
                        pos1 = 0
                        for i in range(0, len(text)):
                            if text[i] == ":":
                                pos = i + 2
                                break
                        for i in range(pos, len(text)):
                            if text[i] == "\n":
                                pos1 = i
                                break
                        text = text[pos:pos1]                               #截取字典文件当前行的翻译内容
                        for row in range(5, ws.max_row + 1):                #遍历excel表中的字符串Id
                            id1 = ws.cell(row, 1).value
                            if id == id1:
                                ws.cell(row, col).value = text              #ID相同时将翻译内容复制到excel
                                break

wb.save('language_03.21.xlsx')