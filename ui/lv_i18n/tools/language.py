#! python3
from openpyxl import load_workbook
import re
import os

os.system("del /f /s /q ..\\*.yml")       # 删除所有字典文件

def read_dic_ver():                       # 读取头文件中字典版本号
    filename = "uires_ver.h"
    major_string = "UI_RES_DICT_MAJOR"
    minor_string = "UI_RES_DICT_MINOR"
    build_string = "UI_RES_DICT_BUILD"
    with open(filename, mode='r', encoding='UTF-8') as file:
        file_lines = file.readlines()
    for line in file_lines:
        major_string_match_result = re.search(major_string, line)
        if major_string_match_result:
            data = line.split(major_string)
            data = data[1].split('\n')
            major = data[0].strip()

        minor_string_match_result = re.search(minor_string, line)
        if minor_string_match_result:
            data = line.split(minor_string)
            data = data[1].split('\n')
            minor = data[0].strip()

        build_string_match_result = re.search(build_string, line)
        if build_string_match_result:
            data = line.split(build_string)
            data = data[1].split('\n')
            build = data[0].strip()
    return major + '.' + minor + '.' + build

wb = load_workbook('language.xlsx')
ws = wb.active
version = read_dic_ver()                                #获取版本号
for col in range(3, ws.max_column + 1):                 #按列遍历表格
    trs = ws.cell(4, col).value;
    if trs != "Y":                                      #判断是否需要生成字典
        continue
    language = ws.cell(1, col).value                    #获取语言名称
    print(language)
    lang = ws.cell(2, col).value                        #获取语言代码
    with open(f'..\\{language}.yml', 'w', encoding='utf-8', newline='') as file:    #打开/创建文件
        file.write(f'{language}:\n  version: {version}\n  lang: {lang}\n')          #写入版本号、语言名称、语言代码
        for row in range(5, ws.max_row + 1):            #按行遍历表格
            file.write(f'  {ws.cell(row, 2).value}: ')  #写入字符串Id
            text = ws.cell(row, col).value
            if text:
                file.write(f'{text}')                   #写入翻译内容
            file.write('\n')

os.chdir("..\\")
print(os.getcwd())
os.system("lv_i18n compile -t *.yml -o .")