#! python3
from openpyxl import load_workbook

wb = load_workbook('language.xlsx')                 #多语言表格
ws = wb.active

wb1 = load_workbook('FD152_多国语言_211012.xlsx')   #翻译稿
ws1 = wb1.active

version = ws.cell(1, 1).value
for row in range(3, ws.max_row + 1):                #按行遍历多语言表格
    zhText = ws.cell(row, 3).value                  #获取中文字符串内容
    for row1 in range(3, ws1.max_row + 1):          #按行遍历翻译稿表格
        text = ws1.cell(row1, 4).value              #获取翻译稿中的中文字符串内容
        if text == zhText:
            #print(ws1.cell(row1, 3).value)
            ws.cell(row, 2).value = ws1.cell(row1, 3).value     #将翻译稿中的字符串id复制到多语言表格
            break

wb.save('language.xlsx')